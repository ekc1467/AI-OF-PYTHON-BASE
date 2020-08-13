# 모듈 로딩 ------------------------------------------------
import openpyxl
# 데이터 변수 선언 ------------------------------------------
filename = "sample.xlsx"
# 엑셀 파일 열기 ---------------------------------------------
book = openpyxl.load_workbook( filename )
# 시트 추출
sheet = book.worksheets[1]
sheetnames = book.sheetnames;
print('sheet =>', sheet.title)
print('sheet.max_row = {}, sheet.max_column={}'.format(sheet.max_row, sheet.max_column))